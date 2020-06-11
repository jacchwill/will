from openpyxl import load_workbook

#物件
class readxls:
    #預設啟用
    def __init__(self,file):
        self.wb = load_workbook(filename=file, read_only=True)
        
    #讀取表格
    def read(self,sheet):
        self.ws = self.wb[sheet]
        return self.ws
    
    #讀取工作表
    def sheetnames(self):
        return self.wb.sheetnames
 